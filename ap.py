import streamlit as st
import pandas as pd
import io
import base64
import os
from datetime import datetime, timedelta

# ================= 1. THIẾT LẬP GIAO DIỆN =================
st.set_page_config(page_title="Tiến độ PTK-Thiên Sơn", page_icon="logothienson.png", layout="wide", initial_sidebar_state="expanded")

# ================= HÀM ĐỌC ẢNH LOCAL LÀM BACKGROUND =================
def get_base64_image(image_path):
    if os.path.exists(image_path):
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    return None

background_image_path = "background.jpg" 
bg_base64 = get_base64_image(background_image_path)

if bg_base64:
    bg_css = f"""
    <style>
        .stApp::before {{
            content: ""; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
            background-image: url("data:image/jpeg;base64,{bg_base64}");
            background-size: cover; background-position: center; background-repeat: no-repeat;
            z-index: -99999;
        }}
        .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {{ background: transparent !important; }}
    </style>
    """
    st.markdown(bg_css, unsafe_allow_html=True)
else:
    st.markdown("""
    <style>
        .stApp::before { 
            content: ""; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
            background-image: url("https://img.freepik.com/free-photo/green-marble-texture-background_23-2150383431.jpg"); 
            background-size: cover; background-position: center; background-repeat: no-repeat;
            z-index: -99999;
        }
        .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)

# ================= CSS TÙY CHỈNH CHUYÊN SÂU =================
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@24,600,1,0" rel="stylesheet" />
<style>
    html, body { overflow: hidden !important; margin: 0 !important; padding: 0 !important; }
    ::-webkit-scrollbar { display: none !important; }
    * { scrollbar-width: none !important; -ms-overflow-style: none !important; box-sizing: border-box !important; }
    header[data-testid="stHeader"] { background: transparent !important; box-shadow: none !important; }
    [data-testid="stHeaderActionElements"], .stAppDeployButton { display: none !important; }
    footer { display: none !important; }

    /* KHOẢNG CÁCH TRÊN DƯỚI 20PX CHO KHUNG BỘ LỌC (SIDEBAR) */
    [data-testid="stSidebar"] { background-color: transparent !important; border: none !important; }
    [data-testid="stSidebarResizer"] { display: none !important; }
    [data-testid="stSidebar"] > div:first-child {
        background-color: rgba(255, 255, 255, 0.35) !important;
        backdrop-filter: blur(20px) !important; -webkit-backdrop-filter: blur(20px) !important;
        border: 1px solid rgba(255, 255, 255, 0.4) !important; border-radius: 24px !important;
        box-shadow: 0 10px 40px rgba(0,0,0,0.15) !important; 
        margin: 20px 0px 20px 20px !important; 
        width: calc(100% - 20px) !important; 
        height: calc(100vh - 40px) !important; 
        overflow: hidden !important; 
    }
    [data-testid="stSidebarUserContent"] { padding: 5px 20px 20px 20px !important; overflow: hidden !important; }
    
    /* HIỆU ỨNG BÓNG ĐỔ CHO LOGO TRONG SIDEBAR */
    [data-testid="stSidebar"] img {
        filter: drop-shadow(0px 10px 15px rgba(0, 0, 0, 0.25)) drop-shadow(0px 4px 6px rgba(0, 0, 0, 0.15)) !important;
        transition: transform 0.3s ease !important;
    }
    [data-testid="stSidebar"] img:hover {
        transform: scale(1.03);
    }

    .sidebar-title { display: flex; align-items: center; gap: 8px; font-size: 13px; font-weight: 700; color: #0f172a; margin-top: 10px; margin-bottom: 2px; text-transform: uppercase; }
    .sidebar-title .material-symbols-rounded { font-size: 18px; color: #198754; }
    
    [data-testid="stSidebar"] div[data-baseweb="select"] > div, [data-testid="stSidebar"] div[data-baseweb="input"] > div {
        background-color: rgba(255, 255, 255, 0.4) !important; backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.6) !important; border-radius: 12px !important; min-height: 32px !important;
    }
    div.row-widget.stSelectbox, div.row-widget.stMultiSelect { margin-bottom: -5px !important; }

    /* KHOẢNG CÁCH TRÊN DƯỚI 20PX CHO KHUNG CHÍNH */
    .block-container { 
        background-color: rgba(255, 255, 255, 0.35) !important; backdrop-filter: blur(20px) !important; -webkit-backdrop-filter: blur(20px) !important;
        border: 1px solid rgba(255, 255, 255, 0.4) !important; border-radius: 24px !important; box-shadow: 0 10px 40px rgba(0,0,0,0.15) !important;
        width: calc(100% - 40px) !important; max-width: 100% !important; 
        margin: 20px 20px !important; 
        padding: 25px 30px !important; 
        height: calc(100vh - 40px) !important; 
        display: flex !important; flex-direction: column !important; overflow: hidden !important; 
    }
    .block-container * { max-width: 100% !important; }
    .block-container > div[data-testid="stVerticalBlock"] { flex-grow: 1 !important; display: flex !important; flex-direction: column !important; min-height: 0 !important; width: 100% !important; }
    .block-container > div[data-testid="stVerticalBlock"] > div { flex-shrink: 0 !important; width: 100% !important; }
    
    /* BẢNG CHI TIẾT CÔNG VIỆC */
    div.element-container:has([data-testid="stDataFrame"]) { flex-grow: 1 !important; flex-shrink: 1 !important; display: flex !important; flex-direction: column !important; min-height: 0 !important; width: 100% !important; }
    [data-testid="stDataFrame"] { flex-grow: 1 !important; width: 100% !important; min-height: 0 !important; overflow-x: auto !important; }
    [data-testid="stDataFrame"] > div { height: 100% !important; width: 100% !important; overflow: auto !important; }
    [data-testid="stDataFrame"] div::-webkit-scrollbar { width: 8px !important; height: 8px !important; display: block !important; }
    [data-testid="stDataFrame"] div::-webkit-scrollbar-track { background: transparent !important; }
    [data-testid="stDataFrame"] div::-webkit-scrollbar-thumb { background: rgba(15, 23, 42, 0.3) !important; border-radius: 10px !important; }
    [data-testid="stDataFrame"] div::-webkit-scrollbar-thumb:hover { background: rgba(15, 23, 42, 0.5) !important; }

    div[data-testid="stDataFrame"] th { background-color: rgba(226, 232, 240, 0.9) !important; color: #0F172A !important; font-weight: 800 !important; font-size: 14px !important; }
    
    /* CĂN GIỮA TUYỆT ĐỐI KHUNG TIÊU ĐỀ */
    .title-card-center { 
        border-radius: 16px; 
        box-shadow: 0 15px 30px rgba(0, 0, 0, 0.15), 0 5px 10px rgba(0, 0, 0, 0.05); 
        background-color: rgba(255, 255, 255, 0.25) !important; 
        backdrop-filter: blur(15px); -webkit-backdrop-filter: blur(15px); 
        border: none !important; 
        width: fit-content; 
        max-width: 100% !important; 
        margin-left: auto; 
        margin-right: auto; 
        display: flex; 
        align-items: center; 
        justify-content: center; 
        padding: 6px 25px; 
        margin-top: 15px; 
        margin-bottom: 15px;
    }

    .custom-download-link { display: block; float: right; text-align: right; color: #0A3622 !important; font-size: 13.5px !important; font-weight: 700 !important; text-decoration: none !important; margin-top: -30px !important; margin-right: 0px !important; margin-bottom: 5px !important; position: relative; z-index: 9999; cursor: pointer; }
    .custom-download-link:hover { color: #198754 !important; text-decoration: underline !important; }
    
    /* NÚT TẢI EXCEL VÀ LÀM MỚI DỮ LIỆU */
    div[data-testid="stButton"] button {
        background-color: rgba(25, 135, 84, 0.85) !important;
        color: white !important;
        border-radius: 10px !important;
        font-weight: bold !important;
        font-size: 13px !important;
        border: 1px solid rgba(255,255,255,0.5) !important;
        padding: 4px 12px !important;
        min-height: 32px !important;
    }
    div[data-testid="stButton"] button:hover {
        background-color: rgba(25, 135, 84, 1) !important;
        box-shadow: 0 4px 10px rgba(25, 135, 84, 0.3) !important;
    }

    /* 10 Ô KPI */
    .kpi-card { width: 100%; padding: 10px 12px; border-radius: 18px !important; border: 1px solid rgba(255, 255, 255, 0.4) !important; box-shadow: 0 10px 25px rgba(0, 0, 0, 0.15), 0 4px 10px rgba(0, 0, 0, 0.05) !important; display: flex; align-items: center; justify-content: center; position: relative; margin-bottom: 10px; min-height: 85px; background-color: rgba(255, 255, 255, 0.35) !important; backdrop-filter: blur(15px); -webkit-backdrop-filter: blur(15px); transition: transform 0.3s ease, box-shadow 0.3s ease; overflow: hidden !important; }
    .kpi-card:hover { transform: translateY(-3px); box-shadow: 0 15px 30px rgba(0, 0, 0, 0.2), 0 5px 10px rgba(0, 0, 0, 0.1) !important; }
    .kpi-icon-wrapper { position: absolute; left: 12px; width: 60px; height: 60px; border-radius: 15px; display: flex; align-items: center; justify-content: center; z-index: 0; opacity: 0.95; background-color: #cfe2ff !important; color: #0d6efd !important; }
    .kpi-icon-wrapper .material-symbols-rounded { font-size: 48px; color: #0d6efd !important; } 
    .kpi-details { display: flex; flex-direction: column; align-items: center; justify-content: center; z-index: 1; text-align: center; }
    .kpi-title { font-size: 1.1rem; font-weight: 800; opacity: 0.85; text-transform: uppercase; margin-bottom: 2px; white-space: nowrap; color: #0f172a !important; text-shadow: 0px 0px 8px rgba(255, 255, 255, 0.8); }
    .kpi-value { font-size: 2.2rem; font-weight: 900; color: #0f172a !important; line-height: 1.1; text-shadow: 0px 0px 8px rgba(255, 255, 255, 0.8); }
    
    div[data-testid="stRadio"] { width: 100% !important; }
    div[data-testid="stRadio"] > div[role="radiogroup"] { display: flex; flex-direction: row; justify-content: center; gap: 30px; margin-top: 5px; margin-bottom: 15px; flex-wrap: wrap; background-color: transparent !important; }
    div[data-testid="stRadio"] > div[role="radiogroup"] > label { cursor: pointer; background: transparent !important; border: none !important; box-shadow: none !important; padding: 2px !important; }
    div[data-testid="stRadio"] > div[role="radiogroup"] > label p { font-weight: 600 !important; color: #0f172a !important; margin: 0 !important; font-size: 0.95rem !important; }
</style>
""", unsafe_allow_html=True)


# ================= 2. ĐỌC DỮ LIỆU TỪ GOOGLE SHEETS =================
# LƯU Ý: Đã gỡ bỏ @st.cache_data để file luôn cập nhật mới nhất khi tải lại
def load_data():
    sheet_url = "https://docs.google.com/spreadsheets/d/1Ps6Bq1q_asSuR3FW5FXMJ46Tr6G02HWJh3gqX3LGG0M/export?format=csv&gid=162795196"
    try:
        df = pd.read_csv(sheet_url)
    except Exception:
        return pd.DataFrame()
    
    df.columns = df.columns.str.strip()
    
    # -------------------------------------------------------------
    # THUẬT TOÁN NHẬN DIỆN CỘT THÔNG MINH (DỰA TRÊN TỪ KHÓA)
    # -------------------------------------------------------------
    rename_dict = {}
    for col in df.columns:
        c_low = str(col).lower().strip()
        if any(kw in c_low for kw in ['trạng thái', 'tình trạng', 'trình trạng']): 
            rename_dict[col] = 'Tình trạng triển khai'
        elif any(kw in c_low for kw in ['tiến độ', '%']): 
            rename_dict[col] = 'Tiến Độ (%)'
        elif any(kw in c_low for kw in ['đầu việc', 'công việc', 'task']):
            if not any(kw in c_low for kw in ['trạng thái', 'tình trạng', 'tiến độ', 'mức']):
                rename_dict[col] = 'Đầu Việc'
        elif any(kw in c_low for kw in ['hạng mục', 'gói thầu']): 
            rename_dict[col] = 'Hạng Mục'
        elif any(kw in c_low for kw in ['vướng mắc', 'ghi chú', 'ý kiến', 'note', 'lý do']): 
            rename_dict[col] = 'Vướng Mắc'
        elif 'dự án' in c_low and not any(kw in c_low for kw in ['chủ nhiệm', 'cnda', 'quản lý', 'mã']): 
            rename_dict[col] = 'Dự Án'
        elif any(kw in c_low for kw in ['mã dự án', 'mã da']):
            rename_dict[col] = 'Mã Dự Án'
        elif any(kw in c_low for kw in ['chủ nhiệm', 'quản lý', 'trưởng nhóm', 'cnda']): 
            rename_dict[col] = 'Chủ nhiệm dự án'
        elif any(kw in c_low for kw in ['hợp đồng', 'hđ', 'plhđ']): 
            rename_dict[col] = 'Hợp Đồng - PLHĐ'
        elif any(kw in c_low for kw in ['chuyên viên', 'triển khai', 'thực hiện', 'cán bộ', 'nhân sự']): 
            if not any(kw in c_low for kw in ['tình trạng', 'trình trạng', 'trạng thái']):
                rename_dict[col] = 'Chuyên viên thực hiện'
        elif any(kw in c_low for kw in ['bắt đầu', 'start']):
            rename_dict[col] = 'Ngày Bắt Đầu'
        elif any(kw in c_low for kw in ['hoàn thành', 'kết thúc', 'deadline', 'mục tiêu', 'hạn chót']):
            if not any(kw in c_low for kw in ['%', 'tiến độ', 'trạng thái', 'tình trạng']):
                rename_dict[col] = 'Ngày Hoàn Thành'
            
    df.rename(columns=rename_dict, inplace=True)
    df = df.loc[:, ~df.columns.duplicated()]
        
    # FORWARD FILL ĐỂ XỬ LÝ CÁC Ô GỘP (MERGE CELLS) TRONG GOOGLE SHEETS
    cols_to_fill = ['Mã Dự Án', 'Dự Án', 'Hợp Đồng - PLHĐ', 'Hạng Mục', 'Chủ nhiệm dự án', 'Chuyên viên thực hiện']
    for col in cols_to_fill:
        if col in df.columns: 
            df[col] = df[col].replace('', pd.NA).ffill()
            
    df = df.fillna('') 
    
    # XỬ LÝ ĐỊNH DẠNG NGÀY THÁNG LINH HOẠT HƠN
    if 'Ngày Bắt Đầu' in df.columns: 
        df['Ngày_Bat_Dau_Obj'] = pd.to_datetime(df['Ngày Bắt Đầu'].astype(str).str.strip(), dayfirst=True, errors='coerce')
    if 'Ngày Hoàn Thành' in df.columns: 
        df['Ngày_Hoan_Thanh_Obj'] = pd.to_datetime(df['Ngày Hoàn Thành'].astype(str).str.strip(), dayfirst=True, errors='coerce')
        
    return df

# SỬ DỤNG SESSION STATE ĐỂ CHỈ CẬP NHẬT GOOGLE SHEET KHI BẤM F5 HOẶC NÚT LÀM MỚI
if 'raw_data' not in st.session_state:
    with st.spinner("⏳ Đang tải dữ liệu mới nhất từ Google Sheets..."):
        st.session_state.raw_data = load_data()

df = st.session_state.raw_data.copy()

# ================= 3. KHỐI SIDEBAR BỘ LỌC =================
with st.sidebar:
    col_logo1, col_logo2, col_logo3 = st.columns([0.1, 2.8, 0.1])
    with col_logo2:
        st.image("logothienson.png", use_container_width=True) 
        
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    
    # NÚT LÀM MỚI DỮ LIỆU
    if st.button("🔄 LÀM MỚI DỮ LIỆU", use_container_width=True):
        st.session_state.pop('raw_data', None)
        st.rerun()
        
    st.markdown("<h3 style='text-align: left; margin-top: 15px; margin-bottom: 5px; color: #0f172a; font-size: 15px;'>BỘ LỌC DỮ LIỆU</h3>", unsafe_allow_html=True)
    
    unique_projects = [str(p) for p in df['Dự Án'].unique() if pd.notna(p) and str(p).strip() != '' and str(p).lower() != 'nan'] if 'Dự Án' in df.columns else []
    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">domain</span> DỰ ÁN</div>', unsafe_allow_html=True)
    selected_projects = st.multiselect("DỰ ÁN", options=unique_projects, placeholder="Chọn Tất cả", label_visibility="collapsed")

    df_temp = df.copy()
    if selected_projects and 'Dự Án' in df_temp.columns: 
        df_temp = df_temp[df_temp['Dự Án'].astype(str).isin(selected_projects)]

    hd_opts = [str(x) for x in df_temp['Hợp Đồng - PLHĐ'].unique() if pd.notna(x) and str(x).strip() != '' and str(x).lower() != 'nan'] if 'Hợp Đồng - PLHĐ' in df_temp.columns else []
    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">description</span> SỐ HỢP ĐỒNG</div>', unsafe_allow_html=True)
    selected_hd = st.multiselect("SỐ HỢP ĐỒNG", options=hd_opts, placeholder="Chọn Tất cả", label_visibility="collapsed")
    if selected_hd and 'Hợp Đồng - PLHĐ' in df_temp.columns: 
        df_temp = df_temp[df_temp['Hợp Đồng - PLHĐ'].astype(str).isin(selected_hd)]

    hm_opts = [str(x) for x in df_temp['Hạng Mục'].unique() if pd.notna(x) and str(x).strip() != '' and str(x).lower() != 'nan'] if 'Hạng Mục' in df_temp.columns else []
    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">folder_open</span> HẠNG MỤC</div>', unsafe_allow_html=True)
    selected_hm = st.multiselect("HẠNG MỤC", options=hm_opts, placeholder="Chọn Tất cả", label_visibility="collapsed")
    if selected_hm and 'Hạng Mục' in df_temp.columns: 
        df_temp = df_temp[df_temp['Hạng Mục'].astype(str).isin(selected_hm)]

    ql_opts = [str(x) for x in df_temp['Chủ nhiệm dự án'].unique() if pd.notna(x) and str(x).strip() != '' and str(x).lower() != 'nan'] if 'Chủ nhiệm dự án' in df_temp.columns else []
    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">manage_accounts</span> TRƯỞNG NHÓM - CNDA</div>', unsafe_allow_html=True)
    selected_ql = st.multiselect("TRƯỞNG NHÓM - CNDA", options=ql_opts, placeholder="Chọn Tất cả", label_visibility="collapsed")

    cb_opts = [str(x) for x in df_temp['Chuyên viên thực hiện'].unique() if pd.notna(x) and str(x).strip() != '' and str(x).lower() != 'nan'] if 'Chuyên viên thực hiện' in df_temp.columns else []
    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">engineering</span> CHUYÊN VIÊN THỰC HIỆN</div>', unsafe_allow_html=True)
    selected_cb = st.multiselect("CHUYÊN VIÊN THỰC HIỆN", options=cb_opts, placeholder="Chọn Tất cả", label_visibility="collapsed")

    st.markdown('<div class="sidebar-title"><span class="material-symbols-rounded">calendar_month</span> KHOẢNG THỜI GIAN</div>', unsafe_allow_html=True)
    time_filter = st.selectbox("KHOẢNG THỜI GIAN", 
                               ["Tất cả", "Hôm nay", "Tuần này", "Tháng này", "Năm nay", "Tùy chọn khoảng ngày"],
                               label_visibility="collapsed")
    
    today = datetime.today().date()
    start_date, end_date = None, None
    if time_filter == "Hôm nay":
        start_date = end_date = today
    elif time_filter == "Tuần này":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif time_filter == "Tháng này":
        start_date = today.replace(day=1)
        if today.month == 12: end_date = today.replace(year=today.year+1, month=1, day=1) - timedelta(days=1)
        else: end_date = today.replace(month=today.month+1, day=1) - timedelta(days=1)
    elif time_filter == "Năm nay":
        start_date = today.replace(month=1, day=1); end_date = today.replace(month=12, day=31)
    elif time_filter == "Tùy chọn khoảng ngày":
        date_range = st.date_input("Chọn khoảng thời gian:", [today, today])
        if len(date_range) == 2: start_date, end_date = date_range
        elif len(date_range) == 1: start_date = end_date = date_range[0]


# ================= 4. KHỐI CHÍNH (BÊN PHẢI) =================

st.markdown("""
<div class="title-card-center" style="padding: 8px 25px; margin-top: 0px; margin-bottom: 15px;">
    <div style="font-size: 26px; font-weight: 600; color: #0A3622; text-shadow: 0 2px 8px rgba(0,0,0,0.2); margin: 0; padding: 0; line-height: 1.2; text-align: center;">BÁO CÁO CÔNG VIỆC - PHÒNG THIẾT KẾ</div>
</div>
""", unsafe_allow_html=True)

df_display = df.copy()
if start_date and end_date and 'Ngày_Bat_Dau_Obj' in df_display.columns and 'Ngày_Hoan_Thanh_Obj' in df_display.columns:
    start_ts = pd.to_datetime(start_date)
    end_ts = pd.to_datetime(end_date)
    cond_start = df_display['Ngày_Bat_Dau_Obj'].between(start_ts, end_ts)
    cond_end = df_display['Ngày_Hoan_Thanh_Obj'].between(start_ts, end_ts)
    df_display = df_display[cond_start | cond_end]

if selected_projects and 'Dự Án' in df_display.columns: df_display = df_display[df_display['Dự Án'].astype(str).isin(selected_projects)]
if selected_hd and 'Hợp Đồng - PLHĐ' in df_display.columns: df_display = df_display[df_display['Hợp Đồng - PLHĐ'].astype(str).isin(selected_hd)]
if selected_hm and 'Hạng Mục' in df_display.columns: df_display = df_display[df_display['Hạng Mục'].astype(str).isin(selected_hm)]
if selected_ql and 'Chủ nhiệm dự án' in df_display.columns: df_display = df_display[df_display['Chủ nhiệm dự án'].astype(str).isin(selected_ql)]
if selected_cb and 'Chuyên viên thực hiện' in df_display.columns: df_display = df_display[df_display['Chuyên viên thực hiện'].astype(str).isin(selected_cb)]

# --- TÍNH TOÁN 10 CHỈ SỐ KPI ĐẢM BẢO KHÔNG TRÙNG LẶP ---
p_projects = 0
if 'Dự Án' in df_display.columns:
    _prjs = df_display['Dự Án'].astype(str).str.strip().str.upper()
    p_projects = _prjs[(_prjs != '') & (_prjs != 'NAN')].nunique()

p_contracts = 0
if 'Hợp Đồng - PLHĐ' in df_display.columns:
    _cts = df_display['Hợp Đồng - PLHĐ'].astype(str).str.strip().str.upper()
    p_contracts = _cts[(_cts != '') & (_cts != 'NAN')].nunique()

p_categories = 0
if 'Hạng Mục' in df_display.columns:
    _cats = df_display['Hạng Mục'].astype(str).str.strip().str.upper()
    p_categories = _cats[(_cats != '') & (_cats != 'NAN')].nunique()

p_total = len(df_display)

if 'Tiến Độ (%)' in df_display.columns and p_total > 0:
    tien_do_clean = pd.to_numeric(df_display['Tiến Độ (%)'].astype(str).str.replace('%', '', regex=False).str.strip(), errors='coerce')
    p_prog = tien_do_clean.mean()
    if pd.isna(p_prog):
        p_prog = 0
else:
    p_prog = 0

if 'Tình trạng triển khai' in df_display.columns:
    p_done = len(df_display[df_display['Tình trạng triển khai'].astype(str).str.strip() == 'Đã hoàn thành'])
    p_inprogress = len(df_display[df_display['Tình trạng triển khai'].astype(str).str.strip() == 'Đang triển khai'])
    p_notstarted = len(df_display[df_display['Tình trạng triển khai'].astype(str).str.strip() == 'Chưa bắt đầu'])
    p_paused = len(df_display[df_display['Tình trạng triển khai'].astype(str).str.strip() == 'Tạm dừng'])
else:
    p_done = p_inprogress = p_notstarted = p_paused = 0

if 'Vướng Mắc' in df_display.columns:
    _issues = df_display['Vướng Mắc'].astype(str).str.strip().str.lower()
    p_issues = sum((_issues != '') & (_issues != 'nan') & (_issues != 'none') & (_issues != 'null'))
else:
    p_issues = 0

def render_transparent_shadow_kpi(title, value, icon_name):
    return f"""
    <div class="kpi-card">
        <div class="kpi-icon-wrapper">
            <span class="material-symbols-rounded">{icon_name}</span>
        </div>
        <div class="kpi-details">
            <div class="kpi-title">{title}</div>
            <div class="kpi-value">{value}</div>
        </div>
    </div>
    """

# --- HÀNG 1: 5 Ô ---
r1_c1, r1_c2, r1_c3, r1_c4, r1_c5 = st.columns(5)
with r1_c1: st.markdown(render_transparent_shadow_kpi("Tổng Dự án", p_projects, "domain"), unsafe_allow_html=True)
with r1_c2: st.markdown(render_transparent_shadow_kpi("Hợp đồng", p_contracts, "description"), unsafe_allow_html=True)
with r1_c3: st.markdown(render_transparent_shadow_kpi("Hạng mục", p_categories, "folder_open"), unsafe_allow_html=True)
with r1_c4: st.markdown(render_transparent_shadow_kpi("Công việc", p_total, "assignment"), unsafe_allow_html=True)
with r1_c5: st.markdown(render_transparent_shadow_kpi("Tiến độ TB", f"{p_prog:.1f}%", "speed"), unsafe_allow_html=True)

# --- HÀNG 2: 5 Ô ---
r2_c1, r2_c2, r2_c3, r2_c4, r2_c5 = st.columns(5)
with r2_c1: st.markdown(render_transparent_shadow_kpi("Đã hoàn thành", p_done, "check_circle"), unsafe_allow_html=True)
with r2_c2: st.markdown(render_transparent_shadow_kpi("Đang triển khai", p_inprogress, "sync"), unsafe_allow_html=True)
with r2_c3: st.markdown(render_transparent_shadow_kpi("Chưa bắt đầu", p_notstarted, "hourglass_empty"), unsafe_allow_html=True)
with r2_c4: st.markdown(render_transparent_shadow_kpi("Tạm dừng", p_paused, "pause_circle"), unsafe_allow_html=True)
with r2_c5: st.markdown(render_transparent_shadow_kpi("Vướng mắc", p_issues, "error"), unsafe_allow_html=True)

status_options = ["Tất cả", "Chưa bắt đầu", "Đang triển khai", "Đã hoàn thành", "Tạm dừng", "Vướng mắc"]
actual_status = st.radio("Bộ lọc Trạng Thái", options=status_options, horizontal=True, label_visibility="collapsed")

if actual_status != "Tất cả":
    if actual_status == "Vướng mắc":
        if 'Vướng Mắc' in df_display.columns:
            _iss = df_display['Vướng Mắc'].astype(str).str.strip().str.lower()
            df_display = df_display[(_iss != '') & (_iss != 'nan') & (_iss != 'none') & (_iss != 'null')]
        else:
            df_display = df_display.iloc[0:0]
    else:
        df_display = df_display[df_display.get('Tình trạng triển khai', '') == actual_status]

df_export = df_display.copy()
for col in ['Ngày_Bat_Dau_Obj', 'Ngày_Hoan_Thanh_Obj']:
    if col in df_display.columns: df_display = df_display.drop(columns=[col])
    if col in df_export.columns: df_export = df_export.drop(columns=[col])

# TIÊU ĐỀ BẢNG CHI TIẾT
st.markdown("""
<div class="title-card-center">
    <div style="font-size: 18px; font-weight: 600; color: #0A3622; text-shadow: 0 2px 6px rgba(0,0,0,0.15); margin: 0; padding: 0; line-height: 1.2; text-align: center;">BẢNG CHI TIẾT CÔNG VIÊC</div>
</div>
""", unsafe_allow_html=True)

def generate_excel_with_colors(df_data):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_data.to_excel(writer, index=False, sheet_name='TienDo')
    output.seek(0)
    import openpyxl
    from openpyxl.styles import PatternFill
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    green_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    red_fill = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")
    gray_fill = PatternFill(start_color="ADB5BD", end_color="ADB5BD", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEeba", end_color="FFEeba", fill_type="solid")

    status_col_idx = None
    vướng_col_idx = None
    for col_idx, col_name in enumerate(df_data.columns, 1):
        if col_name == 'Tình trạng triển khai': status_col_idx = col_idx
        if col_name == 'Vướng Mắc': vướng_col_idx = col_idx

    for row_idx in range(2, ws.max_row + 1):
        fill_to_use = None
        
        if vướng_col_idx:
            v_val = str(ws.cell(row=row_idx, column=vướng_col_idx).value).strip().lower()
            if v_val and v_val not in ['nan', 'none', '']:
                fill_to_use = yellow_fill

        if not fill_to_use and status_col_idx:
            status_val = str(ws.cell(row=row_idx, column=status_col_idx).value).strip()
            if status_val == 'Đã hoàn thành': fill_to_use = green_fill
            elif status_val == 'Tạm dừng': fill_to_use = red_fill
            elif status_val == 'Chưa bắt đầu': fill_to_use = gray_fill

        if fill_to_use:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill_to_use

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()

try:
    excel_bytes = generate_excel_with_colors(df_export)
    b64 = base64.b64encode(excel_bytes).decode()
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "Bao_cao_tien_do_Thien_Son.xlsx"
except Exception:
    csv_bytes = df_export.to_csv(index=False).encode('utf-8-sig')
    b64 = base64.b64encode(csv_bytes).decode()
    mime_type = "text/csv"
    filename = "Bao_cao_tien_do_Thien_Son.csv"

# NÚT TẢI EXCEL
download_html = f'<a class="custom-download-link" href="data:{mime_type};base64,{b64}" download="{filename}">Tải Excel</a>'
st.markdown(download_html, unsafe_allow_html=True)

priority_map = {'Chưa bắt đầu': 1, 'Đang triển khai': 2, 'Đã hoàn thành': 3, 'Tạm dừng': 4}

if 'Tình trạng triển khai' in df_display.columns and 'Tiến Độ (%)' in df_display.columns:
    df_display['Mức Ưu Tiên'] = df_display['Tình trạng triển khai'].map(priority_map).fillna(99)
    sort_cols = ['Mức Ưu Tiên', 'Tiến Độ (%)']
    if 'Hạng Mục' in df_display.columns: sort_cols = ['Hạng Mục'] + sort_cols
    df_display = df_display.sort_values(by=sort_cols, ascending=[True] * len(sort_cols))
    df_display = df_display.drop(columns=['Mức Ưu Tiên'])

def color_rows(row):
    vm = str(row.get('Vướng Mắc', '')).strip().lower()
    if vm and vm not in ['nan', 'none', '']:
        return ['background-color: #FFEeba; color: #000;'] * len(row)
        
    status = row.get('Tình trạng triển khai', '')
    if status == 'Đã hoàn thành': return ['background-color: #a5d6a7; color: #000;'] * len(row)
    if status == 'Tạm dừng': return ['background-color: #ef9a9a; color: #000;'] * len(row)
    if status == 'Chưa bắt đầu': return ['background-color: #adb5bd; color: #000;'] * len(row)
    return ['background-color: #ffffff; color: #000;'] * len(row)

styled_df = df_display.style.apply(color_rows, axis=1).set_table_styles([{
    'selector': 'th',
    'props': [('background-color', 'rgba(226, 232, 240, 0.9)'), ('color', '#0F172A'), ('font-weight', 'bold')]
}])

st.dataframe(styled_df, use_container_width=True, hide_index=True)
